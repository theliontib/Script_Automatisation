import xlwings as xw
from openpyxl import load_workbook


# ✅ trouver le prochain numéro après 0044
def generer_num_apres_44(ws):

    last_row = ws.range("A" + str(ws.cells.last_cell.row)).end("up").row

    for i in range(2, last_row + 1):

        valeur = ws.range(f"A{i}").value

        if valeur == "RA-CQ-0044":
            return 45  # on commence juste après

    return 45  # fallback



def main():

    print("🚀 Lancement script xlwings")

    # ✅ ouvrir fichier Excel (xlwings)
    wb = xw.Book("Inventaire_SI_test.xlsx")

    ws_casque = wb.sheets["Casque"]
    ws_collabs = wb.sheets["Collabs"]

    # ✅ ouvrir source (lecture)
    wb_source = load_workbook("Export de Périphérique.xlsx")
    ws_source = wb_source.active

    # ✅ trouver ligne après 0044
    ligne_depart = None

    last_row_casque = ws_casque.range("A" + str(ws_casque.cells.last_cell.row)).end("up").row

    for i in range(2, last_row_casque + 1):

        if ws_casque.range(f"A{i}").value == "RA-CQ-0044":
            ligne_depart = i + 1
            break

    if ligne_depart is None:
        print("❌ 0044 non trouvé")
        return

    print(f"✅ insertion après ligne {ligne_depart - 1}")

    # ✅ index collaborateurs
    collab_index = {}

    last_row_collab = ws_collabs.range("A" + str(ws_collabs.cells.last_cell.row)).end("up").row

    for i in range(2, last_row_collab + 1):

        nom = str(ws_collabs.range(f"A{i}").value).strip().lower()
        prenom = str(ws_collabs.range(f"B{i}").value).strip().lower()

        key = f"{nom}_{prenom}"

        collab_index[key] = i

    # ✅ récupération numéro UNE seule fois
    numero_courant = generer_num_apres_44(ws_casque)

    ligne_courante = ligne_depart

    # ✅ boucle principale
    for row in ws_source.iter_rows(min_row=2, values_only=True):

        serial = str(row[6]).strip() if row[6] else ""
        nom_complet = str(row[7]).strip() if row[7] else ""

        if not serial:
            continue

        # ✅ date propre
        if row[8]:
            date_achat = row[8]
        else:
            date_achat = None

        # ✅ split nom prénom
        parts = nom_complet.lower().split()

        if len(parts) < 2:
            print(f"⚠️ Nom invalide : {nom_complet}")
            continue

        prenom = parts[0]
        nom = parts[-1]

        key = f"{nom}_{prenom}"

        # ✅ numéro inventaire
        num_inventaire = f"RA-CQ-{numero_courant:04d}"
        numero_courant += 1

        print(f"✅ Ajout casque {num_inventaire}")

        # ✅ écrire dans feuille Casque
        ws_casque.range(f"A{ligne_courante}").value = num_inventaire
        ws_casque.range(f"D{ligne_courante}").value = serial
        ws_casque.range(f"E{ligne_courante}").value = date_achat

        # ✅ assignation collab (menu déroulant conservé)
        if key in collab_index:

            ligne_col = collab_index[key]

            cell = ws_collabs.range(f"L{ligne_col}")

            if not cell.value:
                cell.value = num_inventaire
                print(f"✅ assigné à {nom_complet}")
            else:
                print(f"⛔ déjà attribué : {nom_complet}")

        else:
            print(f"❌ non trouvé : {nom_complet}")

        ligne_courante += 1

    wb.save()

    print("✅ TERMINÉ (propre + menu OK)")


if __name__ == "__main__":
    main()
